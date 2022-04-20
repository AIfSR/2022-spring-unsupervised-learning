from AIfSR_Trajectory_Analysis.output_results.OutputResultsBase import OutputResultsBase
import openpyxl

class OutputXlsx (OutputResultsBase):
    PREDICTION_THRESHOLD = 0.5
    def __init__(self, locationOfXlsx:str, sheetName:str, labels:list[str]) -> None:
        self._locationOfXlsx = locationOfXlsx
        self._sheetName = sheetName
        self._labels = labels

    def checklocationOfXlsx(location:str) -> bool:
        try:
            openpyxl.load_workbook(location)
            return True
        except Exception as e:
            return False
        
    def _addHeaderInfo(self, sheet:openpyxl.worksheet.worksheet, results:list[tuple[str, list[float]]]):
        sheet.append(["Label", "Percentage"])
        def getPercentageOfIndex(i:int):
            totalCount = 0
            iCount = 0
            for _, prediction in results:
                if(prediction[i] > OutputXlsx.PREDICTION_THRESHOLD):
                    iCount +=1
                totalCount += 1
            return iCount / totalCount

        for i in range(len(self._labels)):
            sheet.append([self._labels[i], getPercentageOfIndex(i)])
        sheet.append([])

    def _addIndividualInfo(self, sheet:openpyxl.worksheet.worksheet, results:list[tuple[str, list[float]]]):
        header = ["TrajectoryName"]
        for label in self._labels:
            header.append(label)
        sheet.append(header)
        for name, prediction in results:
            row = [name]
            row.extend(prediction)
            sheet.append(row)

    def output(self, results:list[tuple[str, list[float]]]):
        wb = openpyxl.load_workbook(self._locationOfXlsx)
        sheet = wb.create_sheet(self._sheetName)
        self._addHeaderInfo(sheet, results)
        self._addIndividualInfo(sheet, results)
        wb.save(self._locationOfXlsx)